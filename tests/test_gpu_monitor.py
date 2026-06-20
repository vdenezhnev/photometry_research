"""Tests for GPU monitor and master run log."""

from __future__ import annotations

from pathlib import Path

from adaptive_sampling.common.gpu_monitor import GpuSample, summarize_gpu_samples
from adaptive_sampling.sparse_eval.run_log import MASTER_RUN_COLUMNS, append_master_run_log


def test_summarize_gpu_samples() -> None:
    samples = [
        GpuSample(10, 5, 1000, 8000, 55),
        GpuSample(80, 40, 4000, 8000, 70),
        GpuSample(20, 10, 1500, 8000, 60),
    ]
    stats = summarize_gpu_samples(
        samples,
        gpu_name="Test GPU",
        device_requested="cuda",
        monitor_duration_sec=3.0,
        sample_interval_sec=1.0,
        active_util_threshold=5.0,
    )
    assert stats.gpu_available is True
    assert stats.gpu_name == "Test GPU"
    assert stats.utilization_gpu_max == 80.0
    assert stats.utilization_gpu_avg == 36.67
    assert stats.gpu_active_duration_sec == 3.0
    assert stats.memory_used_mb_max == 4000.0


def test_append_master_run_log(tmp_path: Path) -> None:
    path = tmp_path / "all_runs_log.xlsx"
    row = {col: col for col in MASTER_RUN_COLUMNS}
    append_master_run_log(row, path)
    assert path.is_file()

    append_master_run_log({col: f"2_{col}" for col in MASTER_RUN_COLUMNS}, path)
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["runs"]
    assert ws.max_row == 3
