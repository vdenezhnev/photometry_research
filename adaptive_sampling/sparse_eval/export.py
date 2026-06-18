"""Export sparse eval results to Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    Workbook = None  # type: ignore[misc, assignment]
    _IMPORT_ERROR = exc
else:
    _IMPORT_ERROR = None


def _require_openpyxl() -> None:
    if Workbook is None:
        raise ImportError("openpyxl is required: pip install openpyxl") from _IMPORT_ERROR


def _autosize_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max(len("" if c.value is None else str(c.value)) for c in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def write_sparse_run_xlsx(result_dict: dict[str, Any], xlsx_path: Path) -> None:
    _require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    metrics = result_dict.get("metrics") or {}
    ws.append(["key", "value"])
    for key in (
        "video_slug",
        "fps_label",
        "frames_dir",
        "evaluated_at",
        "input_images",
        "registered_images",
        "registered_ratio",
        "sparse_points",
        "mean_track_length",
        "composite_score",
        "passes_criteria",
        "mapper_success",
    ):
        ws.append([key, metrics.get(key) if key in metrics else result_dict.get(key)])
    _autosize_columns(ws)

    durations = result_dict.get("stage_durations_sec") or {}
    if durations:
        ws_st = wb.create_sheet("stages")
        ws_st.append(["stage", "duration_sec"])
        for stage, sec in sorted(durations.items()):
            ws_st.append([stage, sec])
        _autosize_columns(ws_st)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def write_comparison_xlsx(
    *,
    video_slug: str,
    comparison_rows: list[dict[str, Any]],
    top_fps_modes: list[dict[str, Any]],
    xlsx_path: Path,
) -> None:
    _require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "comparison"
    if comparison_rows:
        headers = list(comparison_rows[0].keys())
        ws.append(headers)
        for row in comparison_rows:
            ws.append([row[h] for h in headers])
        _autosize_columns(ws)

    ws_top = wb.create_sheet("top_fps_modes")
    if top_fps_modes:
        headers = list(top_fps_modes[0].keys())
        ws_top.append(headers)
        for row in top_fps_modes:
            ws_top.append([row[h] for h in headers])
        _autosize_columns(ws_top)

    ws_info = wb.create_sheet("info")
    ws_info.append(["video_slug", video_slug])
    ws_info.append(["runs", len(comparison_rows)])
    _autosize_columns(ws_info)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
