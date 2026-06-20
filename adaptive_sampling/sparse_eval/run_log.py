"""Master XLSX log for all sparse eval runs."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

MASTER_RUN_COLUMNS: tuple[str, ...] = (
    "evaluated_at",
    "video_slug",
    "fps_label",
    "output_dir",
    "device_requested",
    "gpu_available",
    "gpu_name",
    "monitor_duration_sec",
    "gpu_active_duration_sec",
    "gpu_active_ratio",
    "gpu_util_avg",
    "gpu_util_max",
    "gpu_mem_util_avg",
    "gpu_mem_util_max",
    "gpu_memory_used_mb_avg",
    "gpu_memory_used_mb_max",
    "gpu_memory_total_mb",
    "gpu_temperature_c_max",
    "gpu_samples_count",
    "total_duration_sec",
    "stage_feature_extractor_sec",
    "stage_feature_matcher_sec",
    "stage_mapper_sec",
    "stage_undistort_sec",
    "stage_patch_match_stereo_sec",
    "stage_stereo_fusion_sec",
    "fused_ply_path",
    "fused_glb_path",
    "fused_point_count",
    "source_frames",
    "eval_frames",
    "input_images",
    "registered_images",
    "registered_ratio",
    "sparse_points",
    "mean_track_length",
    "composite_score",
    "passes_criteria",
    "mapper_success",
    "glb_path",
)


def _autosize(ws) -> None:
    for col_idx, cells in enumerate(ws.columns, start=1):
        max_len = max(len("" if c.value is None else str(c.value)) for c in cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def append_master_run_log(row: dict[str, Any], xlsx_path: Path, *, sheet_name: str = "runs") -> None:
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    values = [row.get(col, "") for col in MASTER_RUN_COLUMNS]

    if xlsx_path.is_file():
        wb = load_workbook(xlsx_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        if ws.max_row == 0:
            ws.append(list(MASTER_RUN_COLUMNS))
        ws.append(values)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(list(MASTER_RUN_COLUMNS))
        ws.append(values)

    _autosize(ws)
    wb.save(xlsx_path)
