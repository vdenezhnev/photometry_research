"""Разметка пар: Excel, подсказки из sparse, соседние кадры."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

from ..paths import resolve_path

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}
_GOOD = frozenset({"good", "хорошая", "ok", "1", "true", "yes", "да"})
_BAD = frozenset({"bad", "плохая", "0", "false", "no", "нет"})


@dataclass(frozen=True)
class PairLabel:
    video_slug: str
    fps_label: str
    frame_a: str
    frame_b: str
    label: int  # 1 good, 0 bad
    suggested_label: str | None = None
    notes: str | None = None
    source: str = "manual"


def normalize_label(raw: Any) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    if s in _GOOD:
        return 1
    if s in _BAD:
        return 0
    raise ValueError(f"Unknown label: {raw!r} (use good/bad or хорошая/плохая)")


def list_frame_files(frames_dir: Path) -> list[Path]:
    frames_dir = resolve_path(frames_dir)
    files = sorted(
        p for p in frames_dir.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES
    )
    return files


def adjacent_pairs(frames_dir: Path) -> Iterator[tuple[str, str]]:
    files = list_frame_files(frames_dir)
    for a, b in zip(files, files[1:]):
        yield a.name, b.name


def load_sparse_metrics_by_run(sparse_root: Path) -> dict[tuple[str, str], dict[str, Any]]:
    """(video_slug, fps_label) → metrics dict."""
    sparse_root = resolve_path(sparse_root)
    out: dict[tuple[str, str], dict[str, Any]] = {}
    for path in sparse_root.glob("*/fps_*/sparse_metrics.json"):
        data = json.loads(path.read_text(encoding="utf-8"))
        key = (data["video_slug"], data["fps_label"])
        out[key] = data.get("metrics") or {}
    return out


def suggest_label_from_sparse(
    metrics: dict[str, Any] | None,
    *,
    good_min_ratio: float,
    require_passes: bool,
) -> str:
    if not metrics:
        return "bad"
    ratio = float(metrics.get("registered_ratio") or 0.0)
    passes = bool(metrics.get("passes_criteria"))
    if require_passes and not passes:
        return "bad"
    if ratio >= good_min_ratio:
        return "good"
    return "bad"


def iter_pairs_for_template(
    frames_root: Path,
    sparse_root: Path,
    *,
    pair_stride: int,
    include_all_on_failed: bool,
    good_min_ratio: float,
    require_passes: bool,
) -> list[dict[str, Any]]:
    frames_root = resolve_path(frames_root)
    sparse_map = load_sparse_metrics_by_run(sparse_root)
    rows: list[dict[str, Any]] = []

    for video_dir in sorted(frames_root.iterdir()):
        if not video_dir.is_dir():
            continue
        video_slug = video_dir.name
        for fps_dir in sorted(video_dir.iterdir()):
            if not fps_dir.is_dir() or not fps_dir.name.startswith("fps_"):
                continue
            fps_label = fps_dir.name
            metrics = sparse_map.get((video_slug, fps_label))
            suggested = suggest_label_from_sparse(
                metrics,
                good_min_ratio=good_min_ratio,
                require_passes=require_passes,
            )
            failed_fps = suggested == "bad"
            pairs = list(adjacent_pairs(fps_dir))
            if not pairs:
                continue
            for i, (fa, fb) in enumerate(pairs):
                if failed_fps and include_all_on_failed:
                    take = True
                else:
                    take = i % max(pair_stride, 1) == 0
                if not take:
                    continue
                rows.append(
                    {
                        "video_slug": video_slug,
                        "fps_label": fps_label,
                        "frame_a": fa,
                        "frame_b": fb,
                        "suggested_label": suggested,
                        "label": "",
                        "notes": "",
                    }
                )
    return rows


def read_manual_xlsx(path: Path) -> list[PairLabel]:
    from openpyxl import load_workbook

    path = resolve_path(path)
    if not path.is_file():
        return []

    wb = load_workbook(path, read_only=True, data_only=True)
    if "pairs" not in wb.sheetnames:
        raise ValueError(f"Sheet 'pairs' not found in {path}")
    ws = wb["pairs"]
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    required = ["video_slug", "fps_label", "frame_a", "frame_b", "label"]
    for r in required:
        if r not in col:
            raise ValueError(f"Column {r!r} missing in pairs sheet")

    out: list[PairLabel] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col["video_slug"]]:
            continue
        label_raw = row[col["label"]]
        label_int = normalize_label(label_raw)
        if label_int is None:
            continue
        notes = row[col["notes"]] if "notes" in col else None
        suggested = row[col["suggested_label"]] if "suggested_label" in col else None
        out.append(
            PairLabel(
                video_slug=str(row[col["video_slug"]]).strip(),
                fps_label=str(row[col["fps_label"]]).strip(),
                frame_a=str(row[col["frame_a"]]).strip(),
                frame_b=str(row[col["frame_b"]]).strip(),
                label=label_int,
                suggested_label=str(suggested).strip() if suggested else None,
                notes=str(notes).strip() if notes else None,
                source="manual",
            )
        )
    wb.close()
    return out


def export_template_xlsx(rows: list[dict[str, Any]], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    path = resolve_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    info = wb.active
    info.title = "readme"
    info["A1"] = "Заполните столбец label: good или bad (хорошая / плохая)"
    info["A1"].font = Font(bold=True)
    info["A2"] = "suggested_label — подсказка из sparse, можно не менять"
    info["A3"] = "Подробнее: data/labels/README.md"

    ws = wb.create_sheet("pairs", 0)
    headers = ["video_slug", "fps_label", "frame_a", "frame_b", "suggested_label", "label", "notes"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(path)


def write_dataset_csv(pairs: list[PairLabel], path: Path) -> None:
    import csv

    path = resolve_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["video_slug", "fps_label", "frame_a", "frame_b", "label", "source", "notes"],
        )
        w.writeheader()
        for p in pairs:
            w.writerow(
                {
                    "video_slug": p.video_slug,
                    "fps_label": p.fps_label,
                    "frame_a": p.frame_a,
                    "frame_b": p.frame_b,
                    "label": p.label,
                    "source": p.source,
                    "notes": p.notes or "",
                }
            )


def pair_image_paths(pair: PairLabel, frames_root: Path) -> tuple[Path, Path]:
    base = resolve_path(frames_root) / pair.video_slug / pair.fps_label
    return base / pair.frame_a, base / pair.frame_b


def validate_pairs_exist(pairs: list[PairLabel], frames_root: Path) -> list[PairLabel]:
    ok: list[PairLabel] = []
    for p in pairs:
        pa, pb = pair_image_paths(p, frames_root)
        if pa.is_file() and pb.is_file():
            ok.append(p)
    return ok
