"""Экспорт отчётов по проверке пар кадров."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .evaluate import FpsPairQualityResult


def _autosize(ws) -> None:
    for col_idx, cells in enumerate(ws.columns, start=1):
        max_len = max(len("" if c.value is None else str(c.value)) for c in cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)


def write_pair_metrics_json(result: FpsPairQualityResult, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(result.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")


def write_summary_json(result: FpsPairQualityResult, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(result.summary_dict(), indent=2, ensure_ascii=False), encoding="utf-8")


def write_problematic_pairs_csv(result: FpsPairQualityResult, path: Path, *, top_k: int = 50) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = sorted(
        result.problematic_pairs,
        key=lambda p: (p.scene_cut_score, -p.similarity, -p.feature_matches),
        reverse=True,
    )[:top_k]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "frame_a",
                "frame_b",
                "status",
                "similarity",
                "feature_matches",
                "scene_cut_score",
                "reasons",
            ],
        )
        w.writeheader()
        for p in rows:
            w.writerow(
                {
                    "frame_a": p.frame_a,
                    "frame_b": p.frame_b,
                    "status": p.status,
                    "similarity": p.similarity,
                    "feature_matches": p.feature_matches,
                    "scene_cut_score": p.scene_cut_score,
                    "reasons": "; ".join(p.reasons),
                }
            )


def write_report_xlsx(result: FpsPairQualityResult, path: Path, *, top_k: int = 50) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_sum = wb.active
    ws_sum.title = "summary"
    ws_sum.append(["key", "value"])
    for k, v in result.summary_dict().items():
        ws_sum.append([k, v])
    _autosize(ws_sum)

    ws_pairs = wb.create_sheet("all_pairs")
    headers = ["frame_a", "frame_b", "status", "similarity", "feature_matches", "scene_cut_score", "reasons"]
    ws_pairs.append(headers)
    for cell in ws_pairs[1]:
        cell.font = Font(bold=True)
    for p in result.pairs:
        ws_pairs.append(
            [
                p.frame_a,
                p.frame_b,
                p.status,
                p.similarity,
                p.feature_matches,
                p.scene_cut_score,
                "; ".join(p.reasons),
            ]
        )
    _autosize(ws_pairs)

    ws_bad = wb.create_sheet("problematic")
    ws_bad.append(headers)
    for cell in ws_bad[1]:
        cell.font = Font(bold=True)
    bad_sorted = sorted(
        result.problematic_pairs,
        key=lambda p: (p.scene_cut_score, -p.similarity),
        reverse=True,
    )[:top_k]
    for p in bad_sorted:
        ws_bad.append(
            [
                p.frame_a,
                p.frame_b,
                p.status,
                p.similarity,
                p.feature_matches,
                p.scene_cut_score,
                "; ".join(p.reasons),
            ]
        )
    _autosize(ws_bad)

    wb.save(path)


def write_comparison_csv(rows: list[dict[str, Any]], path: Path) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)


def write_comparison_xlsx(*, video_slug: str, rows: list[dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "comparison"
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])
        _autosize(ws)
    ws_info = wb.create_sheet("info")
    ws_info.append(["video_slug", video_slug])
    ws_info.append(["fps_modes", len(rows)])
    _autosize(ws_info)
    wb.save(path)
